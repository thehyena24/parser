from openpyxl import *

def WritePPOverall(file, sheet, source, ID):
    wb = load_workbook(file)
    ws = wb[sheet]

    col = source[ID]["code"]

    ws[col + "4"] = source[ID]["data1"]
    ws[col + "5"] = source[ID]["data2"]

    ws[col + "8"] = source[ID]["data3"]
    ws[col + "9"] = source[ID]["data4"]

    ws[col + "16"] = source[ID]["data5"]
    ws[col + "17"] = source[ID]["data6"]
    ws[col + "18"] = source[ID]["data7"]
    ws[col + "19"] = source[ID]["data8"]
    ws[col + "20"] = source[ID]["data9"]
    ws[col + "21"] = source[ID]["data10"]
    ws[col + "22"] = source[ID]["data11"]
    ws[col + "23"] = source[ID]["data12"]
    ws[col + "24"] = source[ID]["data13"]
    ws[col + "25"] = source[ID]["data14"]
    ws[col + "26"] = source[ID]["data15"]
    ws[col + "27"] = source[ID]["data16"]

    ws[col + "31"] = source[ID]["data17"]
    ws[col + "32"] = source[ID]["data18"]
    ws[col + "33"] = source[ID]["data19"]
    ws[col + "34"] = source[ID]["data20"]
    ws[col + "35"] = source[ID]["data21"]
    ws[col + "36"] = source[ID]["data22"]
    ws[col + "37"] = source[ID]["data23"]
    ws[col + "38"] = source[ID]["data24"]

    wb.save(file)
