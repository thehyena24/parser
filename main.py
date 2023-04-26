import os
import re
from reader import *
from writer import *
from data_store import *
from openpyxl import load_workbook


def Overall(dataStore, directory):
    for root, subdirectories, files in os.walk(directory):
        for file in files:
            if re.search('.unique_file_extension', file):
                f = os.path.join(root, file)

                flat_list = []

                reader = Reader(f, flat_list)        #reader is an object; Reader is the class

                i = GetID(f)

                dataStore.total[i]["data1"] += reader.data1()
                dataStore.total[i]["data2"] += reader.data2()
                
                dataStore.total[i]["data3"] += reader.data3()
                dataStore.total[i]["data4"] += reader.data4()
                
                dataStore.total[i]["data5"] += reader.data5()
                dataStore.total[i]["data6"] += reader.data6()
                dataStore.total[i]["data7"] += reader.data7()
                dataStore.total[i]["data8"] += reader.data8()
                dataStore.total[i]["data9"] += reader.data9()
                dataStore.total[i]["data10"] += reader.data10()
                dataStore.total[i]["data11"] += reader.data11()
                dataStore.total[i]["data12"] += reader.data12()

                dataStore.total[i]["data13"] += reader.data13()
                dataStore.total[i]["data14"] += reader.data14()
                dataStore.total[i]["data15"] += reader.data15()
                dataStore.total[i]["data16"] += reader.data16()
                dataStore.total[i]["data17"] += reader.data17()
                dataStore.total[i]["data18"] += reader.data18()
                dataStore.total[i]["data19"] += reader.data19()
                dataStore.total[i]["data20"] += reader.data20()
                dataStore.total[i]["data21"] += reader.data21()
                dataStore.total[i]["data22"] += reader.data22()
                dataStore.total[i]["data23"] += reader.data23()
                dataStore.total[i]["data24"] += reader.data24()



def MonthWise(dataStore, directory, month):
    for root, subdirectories, files in os.walk(directory):
        for file in files:
            if re.search('.day', file):
                f = os.path.join(root, file)

                flat_list = []

                reader = Reader(f, flat_list)        #reader is an object; Reader is the class

                i = GetID(f)
                m = GetMonth(f)

                if m == month:

                    dataStore.total[i]["data1"] += reader.data1()
                    dataStore.total[i]["data2"] += reader.data2()
                    
                    dataStore.total[i]["data3"] += reader.data3()
                    dataStore.total[i]["data4"] += reader.data4()
                    
                    dataStore.total[i]["data5"] += reader.data5()
                    dataStore.total[i]["data6"] += reader.data6()
                    dataStore.total[i]["data7"] += reader.data7()
                    dataStore.total[i]["data8"] += reader.data8()
                    dataStore.total[i]["data9"] += reader.data9()
                    dataStore.total[i]["data10"] += reader.data10()
                    dataStore.total[i]["data11"] += reader.data11()
                    dataStore.total[i]["data12"] += reader.data12()

                    dataStore.total[i]["data13"] += reader.data13()
                    dataStore.total[i]["data14"] += reader.data14()
                    dataStore.total[i]["data15"] += reader.data15()
                    dataStore.total[i]["data16"] += reader.data16()
                    dataStore.total[i]["data17"] += reader.data17()
                    dataStore.total[i]["data18"] += reader.data18()
                    dataStore.total[i]["data19"] += reader.data19()
                    dataStore.total[i]["data20"] += reader.data20()
                    dataStore.total[i]["data21"] += reader.data21()
                    dataStore.total[i]["data22"] += reader.data22()
                    dataStore.total[i]["data23"] += reader.data23()
                    dataStore.total[i]["data24"] += reader.data24()



def ResetDataStore(dataStore):
    for i in range(1, 17):
        dataStore.total[i]["data1"] = 0
        dataStore.total[i]["data2"] = 0
                    
        dataStore.total[i]["data3"] = 0
        dataStore.total[i]["data4"] = 0
                    
        dataStore.total[i]["data5"] = 0
        dataStore.total[i]["data6"] = 0
        dataStore.total[i]["data7"] = 0
        dataStore.total[i]["data8"] = 0
        dataStore.total[i]["data9"] = 0
        dataStore.total[i]["data10"] = 0
        dataStore.total[i]["data11"] = 0
        dataStore.total[i]["data12"] = 0

        dataStore.total[i]["data13"] = 0
        dataStore.total[i]["data14"] = 0
        dataStore.total[i]["data15"] = 0
        dataStore.total[i]["data16"] = 0
        dataStore.total[i]["data17"] = 0
        dataStore.total[i]["data18"] = 0
        dataStore.total[i]["data19"] = 0
        dataStore.total[i]["data20"] = 0
        dataStore.total[i]["data21"] = 0
        dataStore.total[i]["data22"] = 0
        dataStore.total[i]["data23"] = 0
        dataStore.total[i]["data24"] = 0
        
    


def UpdateSheet(workbook, worksheet, dataStore):
    for i in range(1, 17): 
        WritePPOverall(workbook, worksheet, dataStore.total, i)



directory = input("Enter location of directory to be read: ")
workbook = input("Enter location of workbook to be updated: ")
print("\n")
print("Parsing in progress. This may take several minutes...")
print("\n")

data = DataStore()

monthSheetList = ["Jan 2022", "Feb 2022", "Mar 2022", "Apr 2022", "May 2022", "Jun 2022", "Jul 2022", "Aug 2022", "Sep 2022", "Oct 2022", "Nov 2022", "Dec 2022"]
month = 1

wb = load_workbook(workbook)

Overall(data, directory)
UpdateSheet(workbook, "OVERALL", data)
ResetDataStore(data)

for sheet in monthSheetList:

    if sheet in wb.sheetnames:
        MonthWise(data, directory, month)
        UpdateSheet(workbook, sheet, data)
        ResetDataStore(data)
        month += 1

    else:
        print(sheet, " sheet is not present in the workbook. Skipping", sheet)
        month += 1
        pass

print("\n")
print("Parsing complete. Worbook has been updated.")
